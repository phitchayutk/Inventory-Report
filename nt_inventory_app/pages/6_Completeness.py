import streamlit as st
import pandas as pd
from datetime import datetime
import io
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from zone_db_manager import render_zone_db_selector
from session_manager import render_session_manager, init_session
from report_date_widget import render_report_date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Completeness | NT Report", page_icon="🔎", layout="wide")

for key, default in [
    ('inventory_rows',   []),
    ('port_status_rows', []),
    ('wan_link_rows',    []),
    ('version_map',      {}),
]:
    if key not in st.session_state:
        st.session_state[key] = default

render_zone_db_selector(location="sidebar")
render_report_date()
render_session_manager()

st.title("🔎 Completeness Check")
st.caption("ตรวจสอบว่า hostname ใดขาดหายไปจาก section ไหนบ้าง")

# ── Collect unique hostnames per section ───────────────────────────────────────
inv_rows  = st.session_state.inventory_rows
ps_rows   = st.session_state.port_status_rows
wan_rows  = st.session_state.wan_link_rows
ver_map   = st.session_state.version_map

inv_hosts = set(r.get('Hostname','') for r in inv_rows  if r.get('Hostname'))
ps_hosts  = set(r.get('Hostname','') for r in ps_rows   if r.get('Hostname'))
wan_hosts = set(r.get('Source Hostname','') for r in wan_rows if r.get('Source Hostname'))
ver_hosts = set(ver_map.keys())

all_hosts = inv_hosts | ps_hosts | wan_hosts | ver_hosts

if not all_hosts:
    st.warning("⚠️ ยังไม่มีข้อมูล — กรุณา process ข้อมูลจากหน้า Inventory, Port Status, WAN Link, และ SW Version ก่อน")
    st.stop()

# ── Status summary ─────────────────────────────────────────────────────────────
c1,c2,c3,c4 = st.columns(4)
c1.metric("📦 Inventory",  f"{len(inv_hosts):,} hosts",  delta=f"{'✅' if inv_hosts else '❌'}", delta_color="off")
c2.metric("🔌 Port Status",f"{len(ps_hosts):,} hosts",   delta=f"{'✅' if ps_hosts  else '❌'}", delta_color="off")
c3.metric("🔗 WAN Link",   f"{len(wan_hosts):,} hosts",  delta=f"{'✅' if wan_hosts  else '❌'}", delta_color="off")
c4.metric("🖥️ SW Version", f"{len(ver_hosts):,} hosts",  delta=f"{'✅' if ver_hosts  else '❌'}", delta_color="off")

st.divider()

# ── Build completeness matrix ──────────────────────────────────────────────────
records = []
for hn in sorted(all_hosts):
    in_inv  = hn in inv_hosts
    in_ps   = hn in ps_hosts
    in_wan  = hn in wan_hosts
    in_ver  = hn in ver_hosts
    missing = []
    if not in_inv:  missing.append('Inventory')
    if not in_ps:   missing.append('Port Status')
    if not in_wan:  missing.append('WAN Link')
    if not in_ver:  missing.append('SW Version')
    records.append({
        'Hostname':    hn,
        'Inventory':   '✅' if in_inv  else '❌',
        'Port Status': '✅' if in_ps   else '❌',
        'WAN Link':    '✅' if in_wan  else '❌',
        'SW Version':  '✅' if in_ver  else '❌',
        'Missing':     ', '.join(missing) if missing else '-',
        'Complete':    len(missing) == 0,
        '_missing_count': len(missing),
    })

df_all = pd.DataFrame(records)
df_miss = df_all[df_all['_missing_count'] > 0].copy()
df_ok   = df_all[df_all['_missing_count'] == 0].copy()

# ── Summary metrics ────────────────────────────────────────────────────────────
st.subheader("📊 สรุป")
s1,s2,s3 = st.columns(3)
s1.metric("Total Hostnames",  f"{len(df_all):,}")
s2.metric("✅ ครบทุก section", f"{len(df_ok):,}")
s3.metric("❌ ขาด section",   f"{len(df_miss):,}")

# Missing breakdown per section
if not df_miss.empty:
    st.divider()
    st.subheader("❌ Missing per Section")
    b1,b2,b3,b4 = st.columns(4)
    b1.metric("ขาด Inventory",  f"{(df_miss['Inventory']  =='❌').sum():,}")
    b2.metric("ขาด Port Status",f"{(df_miss['Port Status']=='❌').sum():,}")
    b3.metric("ขาด WAN Link",   f"{(df_miss['WAN Link']   =='❌').sum():,}")
    b4.metric("ขาด SW Version", f"{(df_miss['SW Version'] =='❌').sum():,}")

st.divider()

# ── Filter + Table ─────────────────────────────────────────────────────────────
st.subheader("📋 รายละเอียด")

col_filter, col_export = st.columns([3,1])
with col_filter:
    show_opt = st.radio(
        "แสดง",
        ["❌ เฉพาะที่ขาด", "✅ ครบทุก section", "ทั้งหมด"],
        horizontal=True,
        key="completeness_filter",
    )

display_cols = ['Hostname','Inventory','Port Status','WAN Link','SW Version','Missing']

if show_opt == "❌ เฉพาะที่ขาด":
    df_show = df_miss[display_cols]
elif show_opt == "✅ ครบทุก section":
    df_show = df_ok[display_cols]
else:
    df_show = df_all[display_cols]

# Export button
with col_export:
    # Build Excel for completeness report
    def _build_completeness_excel(df: pd.DataFrame) -> bytes:
        wb  = Workbook()
        ws  = wb.active
        ws.title = 'Completeness'

        FONT_NAME = 'Tahoma'
        _THIN     = Side(style='thin')
        BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
        HDR_FILL  = PatternFill('solid', fgColor='4472C4')
        HDR_FONT  = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
        DATA_FONT = Font(name=FONT_NAME, size=10)
        OK_FILL   = PatternFill('solid', fgColor='E2EFDA')
        MISS_FILL = PatternFill('solid', fgColor='FFE0E0')
        CTR       = Alignment(vertical='center', horizontal='center')
        LEFT      = Alignment(vertical='center', horizontal='left')

        title = f'NT Completeness Check — {datetime.now().strftime("%d-%m-%Y")}'
        ws['A1'] = title
        ws['A1'].font = Font(name=FONT_NAME, bold=True, size=10)
        ws.merge_cells(f'A1:{get_column_letter(len(display_cols))}1')
        ws.row_dimensions[1].height = 20

        col_widths = [28, 12, 12, 12, 12, 40]
        for ci, (col, w) in enumerate(zip(display_cols, col_widths), 1):
            cell = ws.cell(row=2, column=ci, value=col)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = CTR; cell.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w

        all_records = df.to_dict('records')
        for ri, row in enumerate(all_records, 3):
            is_miss = row.get('Missing','-') != '-'
            row_fill = MISS_FILL if is_miss else OK_FILL
            for ci, col in enumerate(display_cols, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(col,''))
                cell.font   = DATA_FONT
                cell.fill   = row_fill
                cell.border = BORDER
                cell.alignment = LEFT if ci in (1,6) else CTR

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:{get_column_letter(len(display_cols))}{len(all_records)+2}'

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf.read()

    excel_bytes = _build_completeness_excel(df_show)
    st.download_button(
        "⬇️ Export Completeness",
        data=excel_bytes,
        file_name=f"NT_Completeness_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# Styled dataframe
st.dataframe(
    df_show.style.apply(
        lambda col: [
            'background-color: #FFE0E0' if v == '❌' else
            'background-color: #E2EFDA' if v == '✅' else ''
            for v in col
        ],
        subset=['Inventory','Port Status','WAN Link','SW Version']
    ),
    use_container_width=True,
    height=500,
)
st.caption(f"แสดง {len(df_show):,} จาก {len(df_all):,} hostnames ทั้งหมด")

# ── Missing detail per section ─────────────────────────────────────────────────
if not df_miss.empty:
    st.divider()
    st.subheader("🔍 รายชื่อ Host ที่ขาด แยกตาม Section")

    tabs = st.tabs(["📦 ขาด Inventory","🔌 ขาด Port Status","🔗 ขาด WAN Link","🖥️ ขาด SW Version"])

    sections = ['Inventory','Port Status','WAN Link','SW Version']
    for tab, sec in zip(tabs, sections):
        with tab:
            miss_sec = df_miss[df_miss[sec] == '❌'][['Hostname','Missing']]
            if miss_sec.empty:
                st.success(f"✅ ไม่มี hostname ที่ขาด {sec}")
            else:
                st.warning(f"❌ {len(miss_sec):,} hostname ขาด {sec}")
                st.dataframe(miss_sec, use_container_width=True, height=300)
