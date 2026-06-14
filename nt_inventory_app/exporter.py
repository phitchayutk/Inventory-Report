"""
NT Inventory Excel Exporter — Full version
Sheets: NT Overall, Port Status, WAN Link, Pivot, New Device, Off Device
"""

from __future__ import annotations
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME  = 'Tahoma'
HDR_FILL   = PatternFill('solid', fgColor='4472C4')
HDR_FONT   = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=10)
DATA_FONT  = Font(name=FONT_NAME, size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='center')
LEFT_ALIGN = Alignment(vertical='center', horizontal='left')
CTR_ALIGN  = Alignment(vertical='center', horizontal='center')
_THIN      = Side(style='thin')
BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

SPEED_FILLS = {
    '100G': PatternFill('solid', fgColor='C00000'),
    '40G':  PatternFill('solid', fgColor='FF6600'),
    '10G':  PatternFill('solid', fgColor='0070C0'),
    '1G':   PatternFill('solid', fgColor='00B050'),
}

_RED_HEADER_FILL  = PatternFill('solid', fgColor='C00000')
_PINK_HEADER_FILL = PatternFill('solid', fgColor='FF0000')
_PINK_ROW_FILL    = PatternFill('solid', fgColor='FFE0E0')
_DARKRED_ROW_FILL = PatternFill('solid', fgColor='FFD0D0')


import re as _re

# Characters Excel/openpyxl cannot handle
_ILLEGAL_CHARS = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]')

def _clean(value):
    """Strip illegal XML characters that openpyxl cannot write to Excel."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS.sub('', value).strip()
    return value


def _hdr(cell, value, fill=None):
    cell.value     = value
    cell.font      = HDR_FONT
    cell.fill      = fill or HDR_FILL
    cell.alignment = WRAP_ALIGN
    cell.border    = BORDER


def _data(cell, value, align=LEFT_ALIGN, fill=None):
    cell.value     = _clean(value)
    cell.font      = DATA_FONT
    cell.alignment = align
    cell.border    = BORDER
    if fill:
        cell.fill = fill


def _title(ws, value: str, ncols: int):
    ws['A1'] = value
    ws['A1'].font      = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN
    if ncols > 1:
        ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws.row_dimensions[1].height = 20


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── NT Overall ────────────────────────────────────────────────────────────────

NT_COLS = [
    ('Network', 12), ('Hostname', 22), ('IP Address', 18),
    ('Platform', 12), ('Type', 22), ('ProductID', 28),
    ('CollectedSN', 20), ('Site Name', 22), ('Zone', 10),
    ('SW Version', 14), ('EOS', 20), ('LDOS', 20), ('LDOS_Planning', 24),
]


def write_nt_overall(ws, rows: list[dict], report_date: str):
    _title(ws, f'NT Inventory Report (ข้อมูล ณ วันที่ {report_date})', len(NT_COLS))
    for ci, (hdr, w) in enumerate(NT_COLS, 1):
        _hdr(ws.cell(row=2, column=ci), hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 16
    for ri, row in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 15
        vals = [
            row.get('Network', 'MPLS LPE'), row.get('Hostname', ''),
            row.get('IP Address', ''),       row.get('Platform', ''),
            row.get('Type', ''),             row.get('ProductID', ''),
            row.get('CollectedSN', ''),      row.get('Site Name', ''),
            row.get('Zone', ''),             row.get('SW Version', ''),
            row.get('EOS', 'No Announcement'),
            row.get('LDOS', 'No Announcement'),
            row.get('LDOS_Planning', 'No Announcement'),
        ]
        for ci, v in enumerate(vals, 1):
            _data(ws.cell(row=ri, column=ci), v)
    ws.freeze_panes = 'A3'
    if rows:
        ws.auto_filter.ref = f'A2:{get_column_letter(len(NT_COLS))}{len(rows)+2}'


# ── Port Status ───────────────────────────────────────────────────────────────

PS_COLS = [
    ('Zone', 8), ('Network', 12), ('Hostname', 24), ('Site Name', 22),
    ('100G\nDown', 8), ('100G\nUp', 8), ('100G\nAdmin\nDown', 10), ('100G\nTotal', 8),
    ('10G\nDown',  8), ('10G\nUp',  8), ('10G\nAdmin\nDown',  10), ('10G\nTotal',  8),
    ('1G\nDown',   8), ('1G\nUp',   8), ('1G\nAdmin\nDown',   10), ('1G\nTotal',   8),
    ('40G\nDown',  8), ('40G\nUp',  8), ('40G\nAdmin\nDown',  10), ('40G\nTotal',  8),
]
_PS_SPEEDS = ['100G', '10G', '1G', '40G']


def write_port_status(ws, rows: list[dict], report_date: str):
    _title(ws, f'NT Port Status Report ( ข้อมูล ณ วันที่ {report_date})', len(PS_COLS))
    for ci, (hdr, w) in enumerate(PS_COLS, 1):
        speed = next((s for s in ['100G', '40G', '10G', '1G'] if s in hdr), None)
        _hdr(ws.cell(row=2, column=ci), hdr, fill=SPEED_FILLS.get(speed, HDR_FILL))
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 36
    for ri, row in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 15
        pc = row.get('port_counts', {})
        vals = [
            row.get('Zone', ''), row.get('Network', 'MPLS LPE'),
            row.get('Hostname', ''), row.get('Site Name', ''),
        ]
        for spd in _PS_SPEEDS:
            b = pc.get(spd, {})
            dn, up, adm = b.get('Down', 0), b.get('Up', 0), b.get('Admin Down', 0)
            vals += [dn, up, adm, dn + up + adm]
        for ci, v in enumerate(vals, 1):
            _data(ws.cell(row=ri, column=ci), v,
                  align=CTR_ALIGN if ci > 4 else LEFT_ALIGN)
    ws.freeze_panes = 'A3'
    if rows:
        ws.auto_filter.ref = f'A2:{get_column_letter(len(PS_COLS))}{len(rows)+2}'


# ── WAN Link ──────────────────────────────────────────────────────────────────

WAN_COLS = [
    ('Source Hostname', 24), ('Source Interface', 32),
    ('Destination Hostname', 24), ('Destination Interface', 32),
]


def write_wan_link(ws, rows: list[dict], report_date: str):
    _title(ws, f'NT WAN Link ( ข้อมูล ณ วันที่ {report_date})', len(WAN_COLS))
    for ci, (hdr, w) in enumerate(WAN_COLS, 1):
        _hdr(ws.cell(row=2, column=ci), hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 16
    for ri, row in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 15
        vals = [
            row.get('Source Hostname', ''), row.get('Source Interface', ''),
            row.get('Destination Hostname', ''), row.get('Destination Interface', ''),
        ]
        for ci, v in enumerate(vals, 1):
            _data(ws.cell(row=ri, column=ci), v)
    ws.freeze_panes = 'A3'
    if rows:
        ws.auto_filter.ref = f'A2:{get_column_letter(len(WAN_COLS))}{len(rows)+2}'


# ── Pivot ─────────────────────────────────────────────────────────────────────

PIVOT_COLS = [
    ('Hostname', 24), ('ProductID', 24),
    ('SiteName', 22), ('Site', 14),
]


def write_pivot(ws, rows: list[dict], report_date: str):
    _title(ws, f'NT Pivot ( ข้อมูล ณ วันที่ {report_date})', len(PIVOT_COLS))
    for ci, (hdr, w) in enumerate(PIVOT_COLS, 1):
        _hdr(ws.cell(row=2, column=ci), hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 16
    for ri, row in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 15
        vals = [
            row.get('Hostname', ''), row.get('ProductID', ''),
            row.get('SiteName', ''), row.get('Site', ''),
        ]
        for ci, v in enumerate(vals, 1):
            _data(ws.cell(row=ri, column=ci), v)
    ws.freeze_panes = 'A3'
    if rows:
        ws.auto_filter.ref = f'A2:{get_column_letter(len(PIVOT_COLS))}{len(rows)+2}'


# ── New Device / Off Device ───────────────────────────────────────────────────

NEWOFF_COLS = [
    ('Network', 12), ('Hostname', 22), ('IP Address', 18),
    ('Platform', 12), ('Type', 14), ('ProductID', 24),
    ('CollectedSN', 20), ('Site Name', 22), ('Zone', 10), ('SW Version', 14),
]


def _write_newoff(ws, rows: list[dict], title_text: str,
                  header_fill: PatternFill, row_fill: PatternFill):
    ws['A1'] = title_text
    ws['A1'].font      = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN
    ws.merge_cells(f'A1:{get_column_letter(len(NEWOFF_COLS))}1')
    ws.row_dimensions[1].height = 20

    for ci, (hdr, w) in enumerate(NEWOFF_COLS, 1):
        _hdr(ws.cell(row=2, column=ci), hdr, fill=header_fill)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 16

    for ri, row in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 15
        vals = [
            row.get('Network', 'MPLS LPE'), row.get('Hostname', ''),
            row.get('IP Address', ''),       row.get('Platform', ''),
            row.get('Type', ''),             row.get('ProductID', ''),
            row.get('CollectedSN', ''),      row.get('Site Name', ''),
            row.get('Zone', ''),             row.get('SW Version', ''),
        ]
        for ci, v in enumerate(vals, 1):
            _data(ws.cell(row=ri, column=ci), v, fill=row_fill)

    ws.freeze_panes = 'A3'
    if rows:
        ws.auto_filter.ref = f'A2:{get_column_letter(len(NEWOFF_COLS))}{len(rows)+2}'


def write_new_device(ws, rows: list[dict], report_date: str):
    _write_newoff(ws, rows,
                  title_text=f'อุปกรณ์ขึ้นใหม่ประจำเดือน {report_date}',
                  header_fill=_PINK_HEADER_FILL,
                  row_fill=_PINK_ROW_FILL)


def write_off_device(ws, rows: list[dict], report_date: str):
    _write_newoff(ws, rows,
                  title_text=f'อุปกรณ์เลิกใช้งานประจำเดือน {report_date}',
                  header_fill=_RED_HEADER_FILL,
                  row_fill=_DARKRED_ROW_FILL)


# ── Per-sheet export (for per-page download buttons) ─────────────────────────

def export_sheet_bytes(sheet: str, rows: list[dict], report_date: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    if sheet == 'NT Overall':
        ws.title = 'NT Overall'
        write_nt_overall(ws, rows, report_date)
    elif sheet == 'Port Status':
        ws.title = 'Port Status'
        write_port_status(ws, rows, report_date)
    elif sheet == 'WAN Link':
        ws.title = 'WAN Link'
        write_wan_link(ws, rows, report_date)
    elif sheet == 'Pivot':
        ws.title = 'Pivot'
        write_pivot(ws, rows, report_date)
    return _wb_to_bytes(wb)


# ── Full report export (6 sheets) ─────────────────────────────────────────────

def export_to_bytes_full(
    inventory_rows:   list[dict],
    port_status_rows: list[dict],
    wan_link_rows:    list[dict],
    pivot_rows:       list[dict],
    new_device_rows:  list[dict],
    off_device_rows:  list[dict],
    report_date:      str | None = None,
) -> bytes:
    if not report_date:
        report_date = datetime.now().strftime('%d-%m-%Y')

    wb = Workbook()

    ws1 = wb.active; ws1.title = 'NT Overall'
    write_nt_overall(ws1, inventory_rows, report_date)

    ws2 = wb.create_sheet('Port Status')
    write_port_status(ws2, port_status_rows, report_date)

    ws3 = wb.create_sheet('WAN Link')
    write_wan_link(ws3, wan_link_rows, report_date)

    ws4 = wb.create_sheet('Pivot')
    write_pivot(ws4, pivot_rows, report_date)

    ws5 = wb.create_sheet('New Device')
    write_new_device(ws5, new_device_rows, report_date)

    ws6 = wb.create_sheet('Off Device')
    write_off_device(ws6, off_device_rows, report_date)

    return _wb_to_bytes(wb)


# ── Legacy alias ──────────────────────────────────────────────────────────────

def export_to_bytes(inventory_rows, port_status_rows, wan_link_rows,
                    report_date=None) -> bytes:
    """Backward-compatible 3-sheet export."""
    return export_to_bytes_full(
        inventory_rows, port_status_rows, wan_link_rows,
        [], [], [], report_date
    )
