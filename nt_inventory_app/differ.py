"""
New Device / Off Device / Pivot logic
- Compare current vs previous month NT Overall (CHASSIS rows only)
- Key: CollectedSN
- New Device  = SN in current but NOT in previous
- Off Device  = SN in previous but NOT in current
- Pivot       = CHASSIS rows of current → Hostname, ProductID, SiteName, Site(prefix)
"""

from __future__ import annotations
import io
import re
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Load NT Overall sheet from uploaded Excel (previous month report)
# ---------------------------------------------------------------------------

def load_nt_overall_from_excel(uploaded_file) -> list[dict]:
    """
    Read 'NT Overall' sheet from a previously exported report.
    Returns list of dicts (all rows, header from row 2).
    """
    data = uploaded_file.read() if hasattr(uploaded_file, 'read') else open(uploaded_file, 'rb').read()
    wb   = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    # Find NT Overall sheet (case-insensitive)
    sheet_name = next(
        (s for s in wb.sheetnames if 'overall' in s.lower() or 'nt overall' in s.lower()),
        wb.sheetnames[0]
    )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    # Row 1 = title, Row 2 = headers
    headers = [str(c).strip() if c else f'col{i}' for i, c in enumerate(rows[1])]
    result  = []
    for row in rows[2:]:
        if not any(row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _chassis_rows(rows: list[dict]) -> list[dict]:
    """Filter to CHASSIS type rows only."""
    return [r for r in rows if str(r.get('Type', '')).upper() == 'CHASSIS']


# ---------------------------------------------------------------------------
# New Device / Off Device
# ---------------------------------------------------------------------------

def compute_new_off(
    current_rows:  list[dict],
    previous_rows: list[dict],
) -> tuple[list[dict], list[dict]]:
    """
    Compare CHASSIS rows by CollectedSN.
    Returns (new_device_rows, off_device_rows)
    Each row has keys matching NT Overall columns.
    """
    cur_chassis  = _chassis_rows(current_rows)
    prev_chassis = _chassis_rows(previous_rows)

    cur_sns  = {str(r.get('CollectedSN', '')).strip(): r for r in cur_chassis  if r.get('CollectedSN')}
    prev_sns = {str(r.get('CollectedSN', '')).strip(): r for r in prev_chassis if r.get('CollectedSN')}

    new_sns = set(cur_sns.keys())  - set(prev_sns.keys())
    off_sns = set(prev_sns.keys()) - set(cur_sns.keys())

    new_rows = [cur_sns[sn]  for sn in sorted(new_sns)]
    off_rows = [prev_sns[sn] for sn in sorted(off_sns)]

    return new_rows, off_rows


# ---------------------------------------------------------------------------
# Pivot
# ---------------------------------------------------------------------------

def _site_prefix(hostname: str) -> str:
    """Extract site prefix: first two underscore-segments, e.g. 'acr_acr'."""
    parts = hostname.lower().split('_')
    return '_'.join(parts[:2]) if len(parts) >= 2 else hostname.lower()


def _is_main_chassis_pid(pid: str) -> bool:
    """
    Return True if PID represents the main chassis unit (not LC, SFC, RP, FAN, PWR).
    e.g. ASR-9903 ✅ | ASR-9903-LC ❌ | ASR-9912-SFC110 ❌ | ASR-9900-RP-TR ❌
    """
    p = pid.upper()
    if re.search(r'-LC$|-SFC\d*$|-RP\b|-FAN|-PSU|-PEM', p):
        return False
    return True


def compute_pivot(current_rows: list[dict]) -> list[dict]:
    """
    Build Pivot rows — 1 row per Hostname (main CHASSIS only).
    Priority: prefer main chassis PID (not LC/SFC/RP) with non-empty Platform.
    """
    chassis = _chassis_rows(current_rows)

    # Group by hostname
    by_host: dict[str, list[dict]] = {}
    for r in chassis:
        hn = str(r.get('Hostname', '')).strip()
        if not hn or hn == 'unknown':
            continue
        by_host.setdefault(hn, []).append(r)

    pivot = []
    for hn, rows in sorted(by_host.items()):
        # Pick best row: main chassis PID with platform > main chassis > any
        main = next((r for r in rows if _is_main_chassis_pid(r.get('ProductID','')) and r.get('Platform')), None)
        if main is None:
            main = next((r for r in rows if _is_main_chassis_pid(r.get('ProductID',''))), None)
        if main is None:
            main = rows[0]

        pivot.append({
            'Hostname':  hn,
            'ProductID': str(main.get('ProductID', '')).strip(),
            'SiteName':  str(main.get('Site Name', '')).strip(),
            'Site':      _site_prefix(hn),
        })

    return pivot


# ---------------------------------------------------------------------------
# Thai date utilities
# ---------------------------------------------------------------------------

from datetime import date as _date, datetime as _datetime

_THAI_MONTHS = ['', 'มกราคม','กุมภาพันธ์','มีนาคม','เมษายน','พฤษภาคม','มิถุนายน',
                'กรกฎาคม','สิงหาคม','กันยายน','ตุลาคม','พฤศจิกายน','ธันวาคม']


def to_thai_date(d: _date) -> str:
    """Convert date → Thai format: '16 พฤษภาคม 2026'"""
    return f'{d.day} {_THAI_MONTHS[d.month]} {d.year}'


def today_thai() -> str:
    return to_thai_date(_date.today())


# ---------------------------------------------------------------------------
# EOS / LDOS lookup from previous month NT Overall
# ---------------------------------------------------------------------------

def build_eos_ldos_map(previous_rows: list[dict]) -> dict[str, dict]:
    """
    Build {ProductID: {EOS, LDOS}} from previous month NT Overall rows.
    Uses first non-empty EOS/LDOS found per ProductID.
    """
    mapping: dict[str, dict] = {}
    for r in previous_rows:
        pid = str(r.get('ProductID', '')).strip()
        if not pid or pid in ('N/A', ''):
            continue
        if pid in mapping:
            continue
        eos  = str(r.get('EOS',  '') or '').strip()
        ldos = str(r.get('LDOS', '') or '').strip()
        if eos or ldos:
            mapping[pid] = {
                'EOS':  eos  or 'No Announcement',
                'LDOS': ldos or 'No Announcement',
            }
    return mapping


# ---------------------------------------------------------------------------
# LDOS_Planning calculation
# ---------------------------------------------------------------------------

def _parse_date_flexible(s: str) -> _date | None:
    """Try multiple date formats."""
    for fmt in ('%d/%b/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return _datetime.strptime(s.strip(), fmt).date()
        except Exception:
            continue
    return None


def calc_ldos_planning(ldos_str: str, report_date: _date) -> str:
    """Calculate 'In X Years Y Months' from LDOS date and report date."""
    if not ldos_str or ldos_str in ('No Announcement', '', 'N/A', 'n/a'):
        return 'No Announcement'

    ldos = _parse_date_flexible(ldos_str)
    if ldos is None:
        return ldos_str  # Return as-is if unparseable

    if ldos <= report_date:
        return 'Expired'

    years  = ldos.year  - report_date.year
    months = ldos.month - report_date.month
    if months < 0:
        years  -= 1
        months += 12

    if years == 0:
        return f'In {months} Months'
    if months == 0:
        return f'In {years} Years'
    return f'In {years} Years {months} Months'


def enrich_eos_ldos(
    inventory_rows: list[dict],
    eos_ldos_map:   dict[str, dict],
    report_date:    _date,
) -> list[dict]:
    """
    Apply EOS/LDOS from map and calculate LDOS_Planning for each inventory row.
    Modifies rows in-place and returns them.
    """
    for r in inventory_rows:
        pid  = str(r.get('ProductID', '')).strip()
        info = eos_ldos_map.get(pid, {})
        eos  = info.get('EOS',  'No Announcement')
        ldos = info.get('LDOS', 'No Announcement')
        r['EOS']           = eos
        r['LDOS']          = ldos
        r['LDOS_Planning'] = calc_ldos_planning(ldos, report_date)
    return inventory_rows
