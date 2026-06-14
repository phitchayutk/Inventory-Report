"""
Lookup utilities for NT Inventory Report
  - Zone DB files  → Site Name + Zone  (keyed on hostname 3-char prefix)
  - show version logs → SW Version per hostname
  - Network classification from hostname suffix
"""

from __future__ import annotations
import re
import io
from pathlib import Path
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Network classification
# ---------------------------------------------------------------------------

def classify_network(hostname: str) -> str:
    h = hostname.lower()
    if re.search(r'(aggpe|_lpe\d*$|_ape\d*$)', h):
        return 'MPLS LPE'
    if re.search(r'(_pe\d*$|_psc\d*$)', h):
        return 'MPLS PE'
    return 'MPLS'


# ---------------------------------------------------------------------------
# Zone DB loader  — supports two file formats
# ---------------------------------------------------------------------------

def load_zone_db(source) -> tuple[str, dict[str, tuple[str, str]]]:
    """
    Load a Zone DB file (uploaded UploadedFile or file path).
    Auto-detects format:
      Format A (Map_Zone.xlsx)  — cols: Hostname(prefix7), Sitename, Zone
      Format B (Zone_New.xlsx)  — cols: Sitename, ตัวย่อ(3-char), Zone New

    Returns (db_name, {abbrev: (site_name, zone)})
    """
    if hasattr(source, 'read'):
        data = source.read()
        name = source.name
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    else:
        name = Path(source).name
        wb = load_workbook(str(source), read_only=True, data_only=True)

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return name, {}

    header = [str(c).strip() if c else '' for c in rows[0]]

    # Detect format by header
    # Format B: first col = Sitename (Thai), second = ตัวย่อ
    if len(header) >= 2 and ('ย่อ' in header[1] or 'abbr' in header[1].lower() or
                              'Zone New' in ' '.join(str(h) for h in rows[0])):
        mapping = _load_format_b(rows[1:])
    else:
        mapping = _load_format_a(rows[1:])

    return name, mapping


def _load_format_a(rows) -> dict[str, tuple[str, str]]:
    """Map_Zone.xlsx: Hostname(prefix), Sitename, Zone"""
    mapping = {}
    for row in rows:
        if not row or not row[0]:
            continue
        prefix   = str(row[0]).strip().lower()          # e.g. 'acr_acr'
        sitename = str(row[1]).strip() if row[1] else ''
        zone     = str(row[2]).strip() if row[2] else ''
        # Extract 3-char abbrev from prefix (first part before _)
        abbrev = prefix.split('_')[0]
        # Also store full prefix key for more precise match
        mapping[prefix] = (sitename, zone)
        if abbrev not in mapping:
            mapping[abbrev] = (sitename, zone)
    return mapping


def _load_format_b(rows) -> dict[str, tuple[str, str]]:
    """Zone_New.xlsx: Sitename, ตัวย่อ, Zone New"""
    mapping = {}
    for row in rows:
        if not row or not row[0]:
            continue
        sitename = str(row[0]).strip() if row[0] else ''
        abbrev   = str(row[1]).strip().lower() if row[1] else ''  # e.g. 'acr'
        zone     = str(row[2]).strip() if row[2] else ''
        if abbrev:
            mapping[abbrev] = (sitename, zone)
    return mapping


# ---------------------------------------------------------------------------
# Lookup: hostname → (site_name, zone)
# ---------------------------------------------------------------------------

def lookup_site_zone(hostname: str, mapping: dict) -> tuple[str, str]:
    """
    Match hostname to (site_name, zone).
    Tries multiple strategies:
      1. Full prefix match (7 chars)  → Format A
      2. 3-char abbrev (first segment) → Format B
      3. Second segment 3-char        → e.g. 'atg' from 'xxx_atg_xxx'
    """
    h = hostname.lower()
    parts = h.split('_')

    # Strategy 1: first two segments joined (e.g. 'acr_acr')
    if len(parts) >= 2:
        key = '_'.join(parts[:2])
        if key in mapping:
            return mapping[key]

    # Strategy 2: first segment 3-char abbrev (e.g. 'acr' from 'acr_acr_aggpe1')
    if parts[0] in mapping:
        return mapping[parts[0]]

    # Strategy 3: second segment (e.g. 'kkm' from 'xxx_kkm_yyy')
    if len(parts) >= 2 and parts[1] in mapping:
        return mapping[parts[1]]

    # Strategy 4: 7-char prefix
    if h[:7] in mapping:
        return mapping[h[:7]]

    return ('', '')


# ---------------------------------------------------------------------------
# SW Version parser
# ---------------------------------------------------------------------------

def parse_show_version(text: str) -> tuple[str, str]:
    """Returns (hostname, version_string)"""
    m = re.search(r'(?:CPU0:|^)([a-zA-Z0-9][a-zA-Z0-9_\-]+)#', text, re.MULTILINE)
    hostname = m.group(1) if m else 'unknown'

    version = ''
    # IOS XR
    m = re.search(r'Cisco IOS XR Software.*?Version\s+([\d\.]+)', text, re.IGNORECASE | re.DOTALL)
    if m:
        version = m.group(1)
    else:
        # IOS XE
        m = re.search(r'Cisco IOS.*?Version\s+([\d\.A-Za-z\(\)]+)', text, re.IGNORECASE | re.DOTALL)
        if m:
            version = m.group(1)
        else:
            m = re.search(r'[Vv]ersion[:\s]+([\d\.]+)', text)
            if m:
                version = m.group(1)

    return hostname, version


def build_version_map(files: list[tuple[str, str]]) -> dict[str, str]:
    """Build {hostname: version} from list of (filename, content) tuples."""
    version_map: dict[str, str] = {}
    for fname, content in files:
        try:
            hostname, version = parse_show_version(content)
            if hostname != 'unknown' and version:
                version_map[hostname] = version
        except Exception:
            continue
    return version_map


# ---------------------------------------------------------------------------
# IP + Platform lookup from Inventory rows
# ---------------------------------------------------------------------------

def build_inv_lookup(inventory_rows: list[dict]) -> dict[str, dict]:
    """
    Build {hostname: {IP Address, Platform, Site Name, Zone, Network, SW Version}}
    from inventory rows (use CHASSIS row preferred, else first row).
    """
    lookup: dict[str, dict] = {}
    for r in inventory_rows:
        hn = r.get('Hostname', '')
        if not hn:
            continue
        if hn not in lookup:
            lookup[hn] = {
                'IP Address': r.get('IP Address', ''),
                'Platform':   r.get('Platform', ''),
                'Site Name':  r.get('Site Name', ''),
                'Zone':       r.get('Zone', ''),
                'Network':    r.get('Network', ''),
                'SW Version': r.get('SW Version', ''),
            }
        else:
            # Prefer CHASSIS row for Platform/IP (more reliable)
            if r.get('Type') == 'CHASSIS':
                existing = lookup[hn]
                if r.get('IP Address'): existing['IP Address'] = r['IP Address']
                if r.get('Platform'):   existing['Platform']   = r['Platform']
    return lookup
